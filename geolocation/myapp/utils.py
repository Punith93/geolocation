import openpyxl
import googlemaps
from geolocation.settings import *
import csv

def formatter_address(file):
	gmaps = googlemaps.Client(key='AIzaSyCHfo0-XQOgZHXIainnPbaxNE7IqFQKc7g')
	wb = openpyxl.load_workbook(file)
	sheet = wb.get_sheet_names()[0]
	datasheet = wb.get_sheet_by_name(sheet)
	result = []
	for i in range(1,datasheet.max_row +1):
		data = {}
		address = datasheet.cell(row=i,column =1).value
		geocode_result = gmaps.geocode(address)
		latitude, longitude = geocode_result[0]['geometry']['location']['lat'], \
		geocode_result[0]['geometry']['location']['lng']
		data['Address'] = address
		data['longitude'] = longitude
		data['latitude'] = latitude
		result.append(data)
	status = generate_excel(result)
	return status

def allowed_file(filename):
    ALLOWED_EXTENSIONS = ['txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'zip', 'doc', 'xls', 'csv', 'ods']
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def write_to_excel(path):
	import xlsxwriter

	wb = xlsxwriter.Workbook(path.replace(".csv",".xlsx"))
	ws = wb.add_worksheet("Geo Details")    # your worksheet title here
	with open(path,'r') as csvfile:
	    table = csv.reader(csvfile)
	    i = 0
	    for row in table:
	        ws.write_row(i, 0, row)
	        i += 1
	wb.close()
	

def formatter(res):
	return [
			
			res['Address'],
			res['longitude'], 
			res['latitude']

		]	


def generate_excel(orders):
	if not os.path.isdir(UPLOAD_FOLDER):
			os.mkdir(UPLOAD_FOLDER)
	path = UPLOAD_FOLDER+'/'+'geo_details.csv'
	with open(path, 'w') as csvfile:
		writer = csv.writer(csvfile, delimiter=',', quotechar='"')
		title = ['Address', 'Longitude', 'Latitude']
		writer.writerow(title)
		for order in orders:
			writer.writerow(formatter(order))
	write_to_excel(path)
	return True